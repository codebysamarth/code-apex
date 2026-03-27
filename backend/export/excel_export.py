"""
Excel Export
============
Export BRD to Excel format (.xlsx) for requirements tracking.

Uses openpyxl for workbook generation with multiple sheets:
- Requirements Matrix (all FRs and NFRs)
- Stakeholders
- Decisions
- Timeline
- Gaps

Great for project managers who need to track requirements in spreadsheets.
"""

import io
from datetime import datetime
from typing import Optional


def _check_openpyxl():
    """Check if openpyxl is available."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
        return True
    except ImportError:
        return False


def export_to_excel(brd_data: dict, output_path: Optional[str] = None) -> bytes:
    """
    Export BRD data to Excel format (.xlsx).
    
    Creates a workbook with multiple sheets for different BRD sections.
    Includes conditional formatting and professional styling.
    
    Args:
        brd_data: The BRD extraction result dictionary
        output_path: Optional file path to save workbook (if None, returns bytes)
        
    Returns:
        XLSX content as bytes
    """
    if not _check_openpyxl():
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
    alt_row_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # MoSCoW color coding
    moscow_colors = {
        'Must': PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid'),
        'Should': PatternFill(start_color='FEEBC8', end_color='FEEBC8', fill_type='solid'),
        'Could': PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid'),
        'Wont': PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid'),
    }
    
    def style_header_row(ws, row_num, num_cols):
        """Apply header styling to a row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def auto_width(ws):
        """Auto-adjust column widths based on content."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(50, max(12, max_length + 2))
            ws.column_dimensions[column].width = adjusted_width
    
    # ========================================
    # Sheet 1: Summary
    # ========================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary['A1'] = brd_data.get('project_name', 'Business Requirements Document')
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary['A2'].font = Font(color='718096')
    
    ws_summary['A4'] = "Document Statistics"
    ws_summary['A4'].font = Font(bold=True, size=12)
    
    stats = [
        ("Functional Requirements", len(brd_data.get('functional_reqs', []))),
        ("Non-Functional Requirements", len(brd_data.get('nfrs', []))),
        ("Stakeholders", len(brd_data.get('stakeholders', []))),
        ("Decisions", len(brd_data.get('decisions', []))),
        ("Timeline Milestones", len(brd_data.get('timeline', []))),
        ("Identified Gaps", len(brd_data.get('gaps', []))),
        ("Completeness Score", f"{brd_data.get('completeness_score', 0) * 100:.0f}%"),
    ]
    
    for i, (label, value) in enumerate(stats, start=5):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value
    
    auto_width(ws_summary)
    
    # ========================================
    # Sheet 2: Requirements Matrix
    # ========================================
    ws_reqs = wb.create_sheet("Requirements")
    
    # Headers
    req_headers = ['ID', 'Type', 'Requirement', 'Priority', 'MoSCoW/Category', 'Confidence', 
                   'Source Span', 'Reasoning', 'Status', 'Assigned To', 'Notes']
    for col, header in enumerate(req_headers, start=1):
        ws_reqs.cell(row=1, column=col, value=header)
    style_header_row(ws_reqs, 1, len(req_headers))
    
    row = 2
    
    # Functional requirements
    for req in brd_data.get('functional_reqs', []):
        ws_reqs.cell(row=row, column=1, value=req.get('id', ''))
        ws_reqs.cell(row=row, column=2, value='Functional')
        ws_reqs.cell(row=row, column=3, value=req.get('text', ''))
        ws_reqs.cell(row=row, column=4, value=req.get('priority', ''))
        ws_reqs.cell(row=row, column=5, value=req.get('moscow', ''))
        ws_reqs.cell(row=row, column=6, value=f"{req.get('confidence', 0):.0%}")
        ws_reqs.cell(row=row, column=7, value=req.get('source_span', ''))
        ws_reqs.cell(row=row, column=8, value=req.get('reasoning', ''))
        ws_reqs.cell(row=row, column=9, value='Open')  # Default status
        ws_reqs.cell(row=row, column=10, value='')  # Assigned To (empty for user to fill)
        ws_reqs.cell(row=row, column=11, value='')  # Notes
        
        # Apply MoSCoW color
        moscow = req.get('moscow', '')
        if moscow in moscow_colors:
            ws_reqs.cell(row=row, column=5).fill = moscow_colors[moscow]
        
        # Apply borders and alternating row colors
        for col in range(1, len(req_headers) + 1):
            ws_reqs.cell(row=row, column=col).border = thin_border
            if row % 2 == 0:
                ws_reqs.cell(row=row, column=col).fill = alt_row_fill
        
        row += 1
    
    # Non-functional requirements
    for req in brd_data.get('nfrs', []):
        ws_reqs.cell(row=row, column=1, value=req.get('id', ''))
        ws_reqs.cell(row=row, column=2, value='Non-Functional')
        ws_reqs.cell(row=row, column=3, value=req.get('text', ''))
        ws_reqs.cell(row=row, column=4, value=req.get('priority', 1))
        ws_reqs.cell(row=row, column=5, value=req.get('category', ''))
        ws_reqs.cell(row=row, column=6, value=f"{req.get('confidence', 0):.0%}")
        ws_reqs.cell(row=row, column=7, value='')
        ws_reqs.cell(row=row, column=8, value=req.get('reasoning', ''))
        ws_reqs.cell(row=row, column=9, value='Open')
        ws_reqs.cell(row=row, column=10, value='')
        ws_reqs.cell(row=row, column=11, value='')
        
        for col in range(1, len(req_headers) + 1):
            ws_reqs.cell(row=row, column=col).border = thin_border
            if row % 2 == 0:
                ws_reqs.cell(row=row, column=col).fill = alt_row_fill
        
        row += 1
    
    # Freeze header row
    ws_reqs.freeze_panes = 'A2'
    auto_width(ws_reqs)
    
    # ========================================
    # Sheet 3: Stakeholders
    # ========================================
    ws_sh = wb.create_sheet("Stakeholders")
    
    sh_headers = ['Name', 'Role', 'Influence Score', 'Mentions', 'Contact', 'Notes', 'Reasoning']
    for col, header in enumerate(sh_headers, start=1):
        ws_sh.cell(row=1, column=col, value=header)
    style_header_row(ws_sh, 1, len(sh_headers))
    
    for row, sh in enumerate(brd_data.get('stakeholders', []), start=2):
        ws_sh.cell(row=row, column=1, value=sh.get('name', ''))
        ws_sh.cell(row=row, column=2, value=sh.get('role', ''))
        ws_sh.cell(row=row, column=3, value=f"{sh.get('influence_score', 0):.0%}")
        ws_sh.cell(row=row, column=4, value=sh.get('mentioned_count', 0))
        ws_sh.cell(row=row, column=5, value='')  # Contact - user fills in
        ws_sh.cell(row=row, column=6, value='')  # Notes
        ws_sh.cell(row=row, column=7, value=sh.get('reasoning', ''))
        
        for col in range(1, len(sh_headers) + 1):
            ws_sh.cell(row=row, column=col).border = thin_border
    
    ws_sh.freeze_panes = 'A2'
    auto_width(ws_sh)
    
    # ========================================
    # Sheet 4: Decisions
    # ========================================
    ws_dec = wb.create_sheet("Decisions")
    
    dec_headers = ['ID', 'Decision', 'Rationale', 'Date', 'Made By', 'Status', 'Reasoning']
    for col, header in enumerate(dec_headers, start=1):
        ws_dec.cell(row=1, column=col, value=header)
    style_header_row(ws_dec, 1, len(dec_headers))
    
    for row, dec in enumerate(brd_data.get('decisions', []), start=2):
        ws_dec.cell(row=row, column=1, value=dec.get('id', ''))
        ws_dec.cell(row=row, column=2, value=dec.get('text', ''))
        ws_dec.cell(row=row, column=3, value=dec.get('rationale', ''))
        ws_dec.cell(row=row, column=4, value=dec.get('date_mentioned', ''))
        ws_dec.cell(row=row, column=5, value='')  # Made By
        ws_dec.cell(row=row, column=6, value='Approved')  # Default status
        ws_dec.cell(row=row, column=7, value=dec.get('reasoning', ''))
        
        for col in range(1, len(dec_headers) + 1):
            ws_dec.cell(row=row, column=col).border = thin_border
    
    ws_dec.freeze_panes = 'A2'
    auto_width(ws_dec)
    
    # ========================================
    # Sheet 5: Timeline
    # ========================================
    ws_tl = wb.create_sheet("Timeline")
    
    tl_headers = ['Milestone', 'Date', 'Type', 'Urgency', 'Dependencies', 'Status', 'Notes']
    for col, header in enumerate(tl_headers, start=1):
        ws_tl.cell(row=1, column=col, value=header)
    style_header_row(ws_tl, 1, len(tl_headers))
    
    for row, ms in enumerate(brd_data.get('timeline', []), start=2):
        ws_tl.cell(row=row, column=1, value=ms.get('milestone', ''))
        ws_tl.cell(row=row, column=2, value=ms.get('date', ''))
        ws_tl.cell(row=row, column=3, value=ms.get('type', ''))
        ws_tl.cell(row=row, column=4, value=ms.get('urgency', ''))
        ws_tl.cell(row=row, column=5, value=', '.join(ms.get('dependencies', [])))
        ws_tl.cell(row=row, column=6, value='Planned')
        ws_tl.cell(row=row, column=7, value='')
        
        for col in range(1, len(tl_headers) + 1):
            ws_tl.cell(row=row, column=col).border = thin_border
    
    ws_tl.freeze_panes = 'A2'
    auto_width(ws_tl)
    
    # ========================================
    # Sheet 6: Gaps
    # ========================================
    ws_gaps = wb.create_sheet("Gaps")
    
    gap_headers = ['Gap', 'Severity', 'Suggestion', 'Status', 'Assigned To', 'Resolution']
    for col, header in enumerate(gap_headers, start=1):
        ws_gaps.cell(row=1, column=col, value=header)
    style_header_row(ws_gaps, 1, len(gap_headers))
    
    severity_colors = {
        'critical': PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid'),
        'high': PatternFill(start_color='FEEBC8', end_color='FEEBC8', fill_type='solid'),
        'medium': PatternFill(start_color='FEFCBF', end_color='FEFCBF', fill_type='solid'),
        'low': PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid'),
    }
    
    for row, gap in enumerate(brd_data.get('gaps', []), start=2):
        ws_gaps.cell(row=row, column=1, value=gap.get('gap', ''))
        ws_gaps.cell(row=row, column=2, value=gap.get('severity', ''))
        ws_gaps.cell(row=row, column=3, value=gap.get('suggestion', ''))
        ws_gaps.cell(row=row, column=4, value='Open')
        ws_gaps.cell(row=row, column=5, value='')
        ws_gaps.cell(row=row, column=6, value='')
        
        severity = gap.get('severity', '')
        if severity in severity_colors:
            ws_gaps.cell(row=row, column=2).fill = severity_colors[severity]
        
        for col in range(1, len(gap_headers) + 1):
            ws_gaps.cell(row=row, column=col).border = thin_border
    
    ws_gaps.freeze_panes = 'A2'
    auto_width(ws_gaps)
    
    # ========================================
    # Sheet 7: Scope
    # ========================================
    ws_scope = wb.create_sheet("Scope")
    
    ws_scope['A1'] = "In Scope"
    ws_scope['A1'].font = Font(bold=True, size=12)
    
    scope = brd_data.get('scope', {})
    row = 2
    for item in scope.get('in_scope', []):
        ws_scope.cell(row=row, column=1, value=f"• {item}")
        row += 1
    
    row += 1
    ws_scope.cell(row=row, column=1, value="Out of Scope")
    ws_scope.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    
    for item in scope.get('out_of_scope', []):
        ws_scope.cell(row=row, column=1, value=f"• {item}")
        row += 1
    
    row += 1
    ws_scope.cell(row=row, column=1, value="Assumptions")
    ws_scope.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    
    for item in scope.get('assumptions', []):
        ws_scope.cell(row=row, column=1, value=f"• {item}")
        row += 1
    
    auto_width(ws_scope)
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()
    buffer.close()
    
    if output_path:
        with open(output_path, 'wb') as f:
            f.write(xlsx_bytes)
    
    return xlsx_bytes


if __name__ == "__main__":
    print(f"openpyxl available: {_check_openpyxl()}")
    
    if _check_openpyxl():
        test_data = {
            "project_name": "Test BRD Export",
            "functional_reqs": [
                {"id": "FR-001", "text": "User authentication via SSO", "priority": 1, "moscow": "Must", "confidence": 0.95, "reasoning": "Explicit requirement"},
                {"id": "FR-002", "text": "Payment processing", "priority": 2, "moscow": "Should", "confidence": 0.85},
            ],
            "nfrs": [
                {"id": "NFR-001", "text": "Response time under 200ms", "category": "Performance", "confidence": 0.88},
            ],
            "stakeholders": [
                {"name": "John Smith", "role": "Product Manager", "influence_score": 0.9, "mentioned_count": 5},
            ],
            "decisions": [
                {"id": "DEC-001", "text": "Use PostgreSQL for database", "rationale": "Better JSON support"},
            ],
            "timeline": [
                {"milestone": "MVP Launch", "date": "Q2 2025", "type": "go-live", "urgency": "high"},
            ],
            "gaps": [
                {"gap": "Security requirements not defined", "severity": "critical", "suggestion": "Add security NFRs"},
            ],
            "scope": {
                "in_scope": ["User management", "Payment processing"],
                "out_of_scope": ["Mobile app"],
                "assumptions": ["SSO provider available"],
            },
            "completeness_score": 0.75,
        }
        
        xlsx_bytes = export_to_excel(test_data)
        print(f"Generated XLSX: {len(xlsx_bytes)} bytes")
